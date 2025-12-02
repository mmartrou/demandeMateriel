def to_dict_request(req):
    # Compatibilité tuple/dict/sqlite3.Row pour les demandes
    if isinstance(req, dict):
        return req
    elif hasattr(req, '_fields'):
        return req._asdict()
    elif hasattr(req, 'keys'):
        return dict(req)
    else:
        # Ordre des colonnes : adapter selon la structure de la table
        return {
            'id': req[0],
            'teacher_id': req[1],
            'request_date': req[2],
            'horaire': req[3],
            'class_name': req[4],
            'material_description': req[5],
            'quantity': req[6],
            'selected_materials': req[7] if len(req) > 7 and req[7] else '',
            'computers_needed': req[8] if len(req) > 8 and req[8] else 0,
            'notes': req[9] if len(req) > 9 else '',
            'prepared': req[10] if len(req) > 10 and req[10] else False,
            'modified': req[11] if len(req) > 11 and req[11] else False,
            'group_count': req[12] if len(req) > 12 else 1,
            'material_prof': req[13] if len(req) > 13 else '',
            'request_name': req[14] if len(req) > 14 else '',
            'room_type': req[15] if len(req) > 15 and req[15] else 'Mixte',
            'image_url': req[16] if len(req) > 16 else None,
            'exam': req[17] if len(req) > 17 else False,
            'created_at': req[18] if len(req) > 18 else None,
            'teacher_name': req[-1] if len(req) > 0 else ''
        }
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import database
from datetime import datetime, timedelta
from ortools.sat.python import cp_model


def h_to_min(hstr):
    """Convert hour string (like '9h30') to minutes"""
    if 'h' in hstr:
        h, m = hstr.split('h')
        return int(h) * 60 + int(m if m else 0)
    else:
        # Handle format like '9:30'
        parts = hstr.split(':')
        return int(parts[0]) * 60 + int(parts[1] if len(parts) > 1 else 0)


def duree_par_niveau(niveau):
    """Get duration by level"""
    if niveau in ("Terminale Spécialité", "SI", "Terminale ES", "1ère Spécialité", "AP 2nd"):
        return 110
    elif niveau in ("AP PP", "1ère ES"):
        return 55
    else:
        return 85


def eleves_par_niveau(niveau, enseignant=None):
    """Get number of students by level"""
    if niveau == "2nde":
        return 20
    else:
        return 20


def interval_cours(c):
    """Get interval (start, end) for a course"""
    start = h_to_min(c["horaire"])
    return (start, start + c["duree"])


def calculer_score_salle(besoins, equipements_salle, type_salle, matiere):
    """
    Calcule un score de compatibilité entre les besoins d'un cours et une salle
    """
    try:
        # Convertir les dictionnaires s'ils ne le sont pas déjà
        if not isinstance(besoins, dict):
            besoins = dict(besoins) if hasattr(besoins, 'items') else {}
        if not isinstance(equipements_salle, dict):
            equipements_salle = dict(equipements_salle) if hasattr(equipements_salle, 'items') else {}
        # Analyser les équipements réels de la salle
        room_has_chemistry_equipment = (equipements_salle.get("eviers", 0) > 0 or equipements_salle.get("hotte", 0) > 0 or 
                                       equipements_salle.get("becs_electriques", 0) > 0 or equipements_salle.get("support_filtration", 0) > 0)
        room_has_physics_equipment = (equipements_salle.get("oscilloscopes", 0) > 0 or equipements_salle.get("bancs_optiques", 0) > 0)
        
        # Vérifier si le cours a des besoins spécifiques
        has_equipment_needs = (besoins.get("ordinateurs", 0) > 0 or besoins.get("eviers", 0) > 0 or 
                              besoins.get("hotte", 0) > 0 or besoins.get("bancs_optiques", 0) > 0 or 
                              besoins.get("oscilloscopes", 0) > 0 or besoins.get("becs_electriques", 0) > 0 or 
                              besoins.get("support_filtration", 0) > 0 or besoins.get("imprimante", 0) > 0)
        
        # Score de base selon la matière et les équipements
        score = 0.5  # Score neutre
        
        matiere_lower = matiere.lower()
        if "chimie" in matiere_lower:
            if room_has_chemistry_equipment:
                score = 1.0 if has_equipment_needs else 0.9
            elif type_salle == "chimie":
                score = 0.7 if has_equipment_needs else 0.8
            else:
                score = 0.3 if has_equipment_needs else 0.6
                
        elif "physique" in matiere_lower:
            if room_has_physics_equipment:
                score = 1.0 if has_equipment_needs else 0.9
            elif type_salle == "physique":
                score = 0.7 if has_equipment_needs else 0.8
            else:
                score = 0.3 if has_equipment_needs else 0.6
        else:
            # Autres matières - préférer salles sans équipements spécialisés
            if not room_has_chemistry_equipment and not room_has_physics_equipment:
                score = 0.8
            else:
                score = 0.4
        
        # Pénaliser si les besoins spécifiques ne sont pas satisfaits
        if besoins.get("eviers", 0) > 0 and equipements_salle.get("eviers", 0) == 0:
            score *= 0.1
        if besoins.get("hotte", 0) > 0 and equipements_salle.get("hotte", 0) == 0:
            score *= 0.1
        if besoins.get("oscilloscopes", 0) > 0 and equipements_salle.get("oscilloscopes", 0) == 0:
            score *= 0.1
        if besoins.get("bancs_optiques", 0) > 0 and equipements_salle.get("bancs_optiques", 0) == 0:
            score *= 0.1
        if besoins.get("becs_electriques", 0) > 0 and equipements_salle.get("becs_electriques", 0) == 0:
            score *= 0.1
        if besoins.get("support_filtration", 0) > 0 and equipements_salle.get("support_filtration", 0) == 0:
            score *= 0.1
        if besoins.get("imprimante", 0) > 0 and equipements_salle.get("imprimante", 0) == 0:
            score *= 0.1
            
        return max(0.0, min(1.0, score))
    except Exception as e:
        print(f"Erreur calcul score salle: {e}")
        print(f"Besoins: {besoins}")
        print(f"Equipements salle: {equipements_salle}")
        return 0.5


def extract_material_needs(selected_materials):
    """Extract material needs from selected_materials string with improved parsing"""
    needs = {
        "ordinateurs": 0,
        "eviers": 0,
        "hotte": 0,
        "bancs_optiques": 0,
        "oscilloscopes": 0,
        "becs_electriques": 0,
        "support_filtration": 0,
        "imprimante": 0,
        "examen": 0
    }
    
    if not selected_materials:
        return needs

    # Convert to lowercase for case-insensitive matching
    materials_text = selected_materials.lower()
    
    # Parse materials - handle both new format (with dashes) and legacy format
    materials = []
    
    # Split by lines first (new format with dashes)
    lines = materials_text.split('\n')
    for line in lines:
        # Remove leading dashes and whitespace
        clean_line = line.strip().lstrip('- ').strip()
        if clean_line:
            materials.append(clean_line)
            # Also split by commas within each line for mixed formats
            if ',' in clean_line:
                materials.extend([m.strip() for m in clean_line.split(',') if m.strip()])
    
    # Also add the full text for pattern matching
    materials.append(materials_text)
    
    # Direct material matching (new system with specific materials)
    specific_materials = {
        'éviers': 'eviers',
        'evier': 'eviers',
        'hotte': 'hotte',
        'bancs optiques': 'bancs_optiques',
        'banc optique': 'bancs_optiques',
        'oscilloscopes': 'oscilloscopes', 
        'oscilloscope': 'oscilloscopes',
        'becs électriques': 'becs_electriques',
        'bec électrique': 'becs_electriques',
        'support de filtration': 'support_filtration',
        'imprimante': 'imprimante'
    }
    
    # Check for direct matches first (more accurate)
    for material in materials:
        material_clean = material.strip().lower()
        
        # Direct material name matches
        if material_clean in specific_materials:
            needs[specific_materials[material_clean]] = 1
            continue
        
    
    return needs


def est_C21_disponible(cours_info, c21_slots):
    """Vérifie si la salle C21 est disponible pour ce cours selon les créneaux configurés"""
    if not c21_slots:
        return True  # Si pas d'info de disponibilité, on considère comme disponible
    
    # Convertir l'horaire du cours
    horaire_debut = cours_info.get('horaire', '8:00')
    duree = cours_info.get('duree', 110)  # Durée par défaut en minutes
    
    debut_cours = h_to_min(horaire_debut)
    fin_cours = debut_cours + duree
    
    # Obtenir le jour de la semaine du cours (pour l'instant on assume 'lundi' par défaut)
    jour_cours = cours_info.get('jour', 'lundi').lower()
    
    # Vérifier si le cours peut tenir dans au moins un des créneaux disponibles pour ce jour
    for slot in c21_slots:
        if slot['jour'].lower() == jour_cours:
            debut_dispo = h_to_min(slot['heure_debut'])
            fin_dispo = h_to_min(slot['heure_fin'])
            
            # Le cours doit être entièrement dans la plage de disponibilité
            if debut_cours >= debut_dispo and fin_cours <= fin_dispo:
                return True
    
    return False  # Aucun créneau ne convient

def compatible(salle, besoin, c21_slots=None):
    """Check if a room is compatible with course needs"""
    # Vérification spécifique pour C21
    if salle.get("nom") == "C21" and c21_slots is not None:
        if not est_C21_disponible(besoin, c21_slots):
            return False
    
    # Check if this is a theoretical course (no specific equipment needs)
    has_equipment_needs = (besoin["ordinateurs"] > 0 or besoin["eviers"] > 0 or 
                          besoin["hotte"] > 0 or besoin["bancs_optiques"] > 0 or 
                          besoin["oscilloscopes"] > 0 or besoin["becs_electriques"] > 0 or 
                          besoin["support_filtration"] > 0 or besoin["imprimante"] > 0 or 
                          besoin["examen"] > 0)
    
    # For theoretical courses (no equipment needs), any room type is acceptable
    # This allows flexibility for courses that don't need specific lab equipment
    if not has_equipment_needs:
        # Still need to check capacity
        if besoin["chaises"] > salle["chaises"]:
            return False
        return True
    
    # For courses with specific equipment needs, check subject compatibility
    if besoin["matiere"] == "chimie" and salle["type"] not in ("chimie", "mixte"):
        return False
    if besoin["matiere"] == "physique" and salle["type"] not in ("physique", "mixte"):
        return False
    
    # Check equipment needs
    if besoin["ordinateurs"] > salle["ordinateurs"]:
        return False
    if besoin["eviers"] > salle["eviers"]:
        return False
    if besoin["chaises"] > salle["chaises"]:
        return False
    if besoin["hotte"] > salle["hotte"]:
        return False
    if besoin["bancs_optiques"] > salle["bancs_optiques"]:
        return False
    if besoin["oscilloscopes"] > salle["oscilloscopes"]:
        return False
    if besoin["becs_electriques"] > salle["becs_electriques"]:
        return False
    if besoin["support_filtration"] > salle["support_filtration"]:
        return False
    if besoin["imprimante"] > salle["imprimante"]:
        return False
    if besoin["examen"] > salle["examen"]:
        return False
    
    return True




def h_to_min(hstr):
    """Convertit une heure en format HH:MM ou HHhMM en minutes."""
    try:
        if 'h' in hstr:
            h, m = map(int, hstr.replace("h", ":").split(":"))
        elif ':' in hstr:
            h, m = map(int, hstr.split(":"))
        else:
            h = int(hstr)
            m = 0
        return h * 60 + m
    except:
        return 8 * 60  # Défaut 8h00

def generer_excel_optimise(cours, salles, x, solver, unassigned_courses, date_param=None, custom_room_assignments=None):
    """Génération Excel optimisée avec le solveur CP - Style grille horaire."""
    try:
        print(f"🔧 [Excel] Début génération Excel avec custom_room_assignments type: {type(custom_room_assignments)}")
        print(f"🔧 [Excel] custom_room_assignments value: {custom_room_assignments}")
        
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
        
        wb = Workbook()
        
        # Créneaux horaires personnalisés (comme dans main.py)
        horaires_str = [
            "9h00", "9h30", "10h00", "10h45", "11h15", "11h45", "12h15", "12h45",
            "13h15", "13h45", "14h15", "14h45", "15h15", "15h45", "16h15", "16h45",
            "17h15", "17h45", "18h15"
        ]
        horaires = [h_to_min(h) for h in horaires_str]
        
        # Définition de l'ordre personnalisé des salles
        physique_salles = ["C23", "C25", "C27", "C22", "C24"]
        chimie_salles = ["C32", "C33", "C31"]
        
        # Liste des salles dans l'ordre demandé: PHYSIQUE puis CHIMIE puis C21
        ordered_rooms = physique_salles + chimie_salles + ["C21"]
        
        # Filtrer pour ne garder que les salles qui existent
        salle_list = [room for room in ordered_rooms if room in salles.keys()]
        
        # Ajouter les salles restantes qui ne sont pas dans l'ordre personnalisé (au cas où)
        remaining_rooms = [room for room in salles.keys() if room not in salle_list]
        salle_list.extend(sorted(remaining_rooms))
        
        # Feuille 1: Planning détaillé (techniciens)
        ws1 = wb.active
        ws1.title = "Planning_Techniciens"
        
        # Styles
        bold_font = Font(bold=True, size=16)
        header_font = Font(bold=True, size=36)
        border = Border(left=Side(style='thick'), right=Side(style='thick'), 
                       top=Side(style='thick'), bottom=Side(style='thick'))
        
        # En-tête principal (ligne 1)
        ws1.cell(row=1, column=1, value="Planning")
        ws1.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws1.column_dimensions['A'].width = 12
        
        # Encadrés Physique et Chimie
        col_physique = [2 + salle_list.index(s) for s in physique_salles if s in salle_list]
        col_chimie = [2 + salle_list.index(s) for s in chimie_salles if s in salle_list]
        
        if col_physique and len(col_physique) > 0:
            # Ne fusionner que s'il y a plusieurs colonnes
            if len(col_physique) > 1 and col_physique[-1] > col_physique[0]:
                ws1.merge_cells(start_row=1, start_column=col_physique[0], end_row=1, end_column=col_physique[-1])
            for col in range(col_physique[0], col_physique[-1]+1):
                cell_phys = ws1.cell(row=1, column=col, value="Physique" if col==col_physique[0] else None)
                cell_phys.font = header_font
                cell_phys.alignment = Alignment(horizontal="center", vertical="center")
                cell_phys.border = border
        
        if col_chimie and len(col_chimie) > 0:
            # Ne fusionner que s'il y a plusieurs colonnes
            if len(col_chimie) > 1 and col_chimie[-1] > col_chimie[0]:
                ws1.merge_cells(start_row=1, start_column=col_chimie[0], end_row=1, end_column=col_chimie[-1])
            for col in range(col_chimie[0], col_chimie[-1]+1):
                cell_chim = ws1.cell(row=1, column=col, value="Chimie" if col==col_chimie[0] else None)
                cell_chim.font = header_font
                cell_chim.alignment = Alignment(horizontal="center", vertical="center")
                cell_chim.border = border
        
        # Noms des salles (ligne 2)
        for idx_s, s in enumerate(salle_list):
            col_letter = ws1.cell(row=2, column=2+idx_s).column_letter
            cell = ws1.cell(row=2, column=2+idx_s, value=s)
            ws1.column_dimensions[col_letter].width = 20
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Date (ligne 2, colonne 1) - Utiliser la date du planning, pas la date actuelle
        from datetime import datetime
        # Si date_param n'est pas fournie, utiliser la date actuelle
        if date_param:
            if isinstance(date_param, str):
                planning_date = datetime.strptime(date_param, '%Y-%m-%d')
            else:
                planning_date = date_param
            formatted_date = planning_date.strftime('%d/%m')
        else:
            formatted_date = datetime.now().strftime('%d/%m')
        ws1.cell(row=2, column=1, value=formatted_date)
        ws1.cell(row=2, column=1).font = bold_font
        ws1.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=2, column=1).border = border
        
        # Préparation matrice pour fusion de cellules
        cell_matrix = [[{"content": "", "content_techniciens": "", "content_affichage": "", "merge": False, "merge_len": 1, "matiere": ""} for _ in salle_list] for _ in horaires]
        
        # Remplir la matrice avec les cours assignés
        for i, c in enumerate(cours):
            # Find assigned room
            salle_assignee = None
            
            # Vérifier si on a une assignation personnalisée pour ce cours
            course_id = c.get('id')
            if custom_room_assignments and course_id:
                try:
                    # S'assurer que custom_room_assignments est un dictionnaire
                    if isinstance(custom_room_assignments, str):
                        import json
                        custom_room_assignments = json.loads(custom_room_assignments)
                    
                    if isinstance(custom_room_assignments, dict):
                        assigned_room_name = custom_room_assignments.get(str(course_id))
                        if assigned_room_name:
                            # Trouver la clé de la salle à partir du nom
                            for room_key, room_data in salles.items():
                                if room_data.get('nom') == assigned_room_name or room_key == assigned_room_name:
                                    salle_assignee = room_key
                                    print(f"🎯 Assignation personnalisée: cours {course_id} -> salle {room_key}")
                                    break
                except Exception as e:
                    print(f"⚠️ Erreur traitement assignations personnalisées: {e}")
                    # En cas d'erreur, continuer avec l'assignation normale
            
            # Si pas d'assignation personnalisée, utiliser l'assignation du solver
            if not salle_assignee:
                for s in salle_list:
                    if (i, s) in x and solver.Value(x[(i, s)]) == 1:
                        salle_assignee = s
                        break
            
            if salle_assignee:
                # Parse course time
                horaire_debut = c.get('horaire', '8:00')
                start_time = h_to_min(horaire_debut)
                duree = c.get('duree', 110)  # Durée par défaut
                
                # Find start and end indices in horaires
                idx_start = 0
                idx_end = 1
                
                # Find the closest start time
                for idx, h in enumerate(horaires):
                    if h <= start_time:
                        idx_start = idx
                    if h >= start_time + duree and idx_end == 1:
                        idx_end = idx
                        break
                else:
                    # Si on n'a pas trouvé de fin, utiliser la fin du tableau
                    idx_end = len(horaires)
                
                # S'assurer que idx_end > idx_start
                if idx_end <= idx_start:
                    idx_end = idx_start + 1
                
                merge_len = idx_end - idx_start
                
                # Extraire les équipements de salle demandés depuis selected_materials
                besoins_equipements = extract_material_needs(c.get('selected_materials', ''))
                
                # Build course content for Planning_Techniciens (équipements de salle + titre TP)
                besoins_txt_techniciens = []
                
                # Ajouter seulement les équipements de salle qui sont réellement demandés
                if besoins_equipements.get('eviers', 0) > 0:
                    besoins_txt_techniciens.append("Éviers")
                if besoins_equipements.get('hotte', 0) > 0:
                    besoins_txt_techniciens.append("Hotte")
                if besoins_equipements.get('bancs_optiques', 0) > 0:
                    besoins_txt_techniciens.append("Bancs optiques")
                if besoins_equipements.get('oscilloscopes', 0) > 0:
                    besoins_txt_techniciens.append("Oscilloscopes")
                if besoins_equipements.get('becs_electriques', 0) > 0:
                    besoins_txt_techniciens.append("Becs électriques")
                if besoins_equipements.get('support_filtration', 0) > 0:
                    besoins_txt_techniciens.append("Support de filtration")
                if besoins_equipements.get('imprimante', 0) > 0:
                    besoins_txt_techniciens.append("Imprimante")
                
                # Ajouter le titre du TP (request_name) s'il existe
                titre_tp = c.get('request_name', '')
                if titre_tp and titre_tp.strip():
                    besoins_txt_techniciens.append(titre_tp.strip())
                
                content_techniciens = f"{c.get('enseignant', '')}\n{c.get('niveau', '')}"
                if besoins_txt_techniciens:
                    content_techniciens += "\n" + ", ".join(besoins_txt_techniciens)
                
                # Build course content for Affichage (seulement enseignant + niveau)
                content_affichage = f"{c.get('enseignant', '')}\n{c.get('niveau', '')}"
                
                # Fill matrix avec vérifications d'indices
                try:
                    salle_idx = salle_list.index(salle_assignee)
                except ValueError:
                    continue  # Salle non trouvée, ignorer
                
                # Vérifier que les indices sont dans les limites
                if (idx_start >= 0 and idx_start < len(cell_matrix) and 
                    salle_idx >= 0 and salle_idx < len(salle_list)):
                    
                    # Remplir seulement la première cellule pour éviter les conflits avec les cellules fusionnées
                    if idx_start < len(cell_matrix) and salle_idx < len(cell_matrix[idx_start]):
                        cell_matrix[idx_start][salle_idx]["content_techniciens"] = content_techniciens
                        cell_matrix[idx_start][salle_idx]["content_affichage"] = content_affichage
                        cell_matrix[idx_start][salle_idx]["content"] = content_techniciens  # Par défaut, pour compatibilité
                        cell_matrix[idx_start][salle_idx]["matiere"] = c.get('matiere', 'mixte')
                        
                        # Marquer les autres cellules comme "occupées" mais sans contenu
                        for idx in range(idx_start + 1, min(idx_end, len(cell_matrix))):
                            if idx < len(cell_matrix) and salle_idx < len(cell_matrix[idx]):
                                cell_matrix[idx][salle_idx]["content_techniciens"] = ""  # Cellule occupée mais vide
                                cell_matrix[idx][salle_idx]["content_affichage"] = ""
                                cell_matrix[idx][salle_idx]["content"] = ""  # Cellule occupée mais vide
                                cell_matrix[idx][salle_idx]["matiere"] = c.get('matiere', 'mixte')
                    
                    # Marquer pour fusion seulement si valide
                    if merge_len > 1 and idx_start < len(cell_matrix):
                        cell_matrix[idx_start][salle_idx]["merge"] = True
                        cell_matrix[idx_start][salle_idx]["merge_len"] = merge_len
        
                # Remplissage du tableau avec fusion verticale
        # D'abord, faire toutes les fusions
        merged_ranges = set()
        for idx_h, h in enumerate(horaires):
            for idx_s, s in enumerate(salle_list):
                cell = cell_matrix[idx_h][idx_s]
                excel_row = 3 + idx_h
                excel_col = 2 + idx_s
                
                # Fusion des cellules - seulement pour les cellules de début avec contenu
                if cell["merge"] and cell["content"] and cell["merge_len"] > 1:
                    end_row = excel_row + cell["merge_len"] - 1
                    merge_key = (excel_row, excel_col, end_row)
                    
                    # Vérifier que la fusion est valide et pas déjà faite
                    if (end_row > excel_row and 
                        end_row <= 3 + len(horaires) - 1 and 
                        merge_key not in merged_ranges):
                        
                        try:
                            ws1.merge_cells(
                                start_row=excel_row, start_column=excel_col,
                                end_row=end_row, end_column=excel_col
                            )
                            merged_ranges.add(merge_key)
                        except Exception as e:
                            print(f"Erreur fusion: start_row={excel_row}, end_row={end_row}, col={excel_col}: {e}")
        
        # Ensuite, remplir le contenu
        for idx_h, h in enumerate(horaires):
            # Horaires (colonne 1)
            ws1.cell(row=3+idx_h, column=1, value=horaires_str[idx_h])
            ws1.cell(row=3+idx_h, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws1.cell(row=3+idx_h, column=1).border = border
            
            # Salles
            for idx_s, s in enumerate(salle_list):
                cell = cell_matrix[idx_h][idx_s]
                excel_row = 3 + idx_h
                excel_col = 2 + idx_s
                
                # Vérifier si cette cellule est une cellule fusionnée secondaire
                is_merged_secondary = False
                for merge_key in merged_ranges:
                    start_row, start_col, end_row = merge_key
                    if (start_col == excel_col and 
                        start_row < excel_row <= end_row):
                        is_merged_secondary = True
                        break
                
                # Appliquer les bordures à toutes les cellules (y compris fusionnées)
                try:
                    ws1.cell(row=excel_row, column=excel_col).border = border
                except Exception:
                    pass
                
                # Remplir seulement les cellules principales (pas les secondaires)
                if not is_merged_secondary:
                    try:
                        # Utiliser le contenu techniciens pour la feuille Planning_Techniciens
                        ws1.cell(row=excel_row, column=excel_col, value=cell["content_techniciens"])
                        ws1.cell(row=excel_row, column=excel_col).alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )
                        
                        # Couleurs seulement si contenu
                        if cell["content"]:
                            matiere = cell["matiere"].lower() if cell["matiere"] else ""
                            if "chimie" in matiere:
                                ws1.cell(row=excel_row, column=excel_col).fill = PatternFill(
                                    start_color="B2F2E9", end_color="B2F2E9", fill_type="solid"
                                )
                            elif "physique" in matiere:
                                ws1.cell(row=excel_row, column=excel_col).fill = PatternFill(
                                    start_color="FFD580", end_color="FFD580", fill_type="solid"
                                )
                            else:
                                ws1.cell(row=excel_row, column=excel_col).fill = PatternFill(
                                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                                )
                    except Exception as e:
                        # Ignorer les erreurs sur les cellules fusionnées
                        print(f"Ignoring cell write error at ({excel_row}, {excel_col}): {e}")
                        pass        # Hauteur des lignes
        for idx_h in range(len(horaires)):
            ws1.row_dimensions[3 + idx_h].height = 20
        
        # Feuille 2: Affichage simplifié
        ws2 = wb.create_sheet(title="Affichage")
        
        # Même structure mais contenu simplifié
        ws2.cell(row=1, column=1, value="Planning")
        ws2.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions['A'].width = 12
        
        # Répliquer les en-têtes
        if col_physique and len(col_physique) > 0:
            # Ne fusionner que s'il y a plusieurs colonnes
            if len(col_physique) > 1 and col_physique[-1] > col_physique[0]:
                ws2.merge_cells(start_row=1, start_column=col_physique[0], end_row=1, end_column=col_physique[-1])
            for col in range(col_physique[0], col_physique[-1]+1):
                cell_phys = ws2.cell(row=1, column=col, value="Physique" if col==col_physique[0] else None)
                cell_phys.font = header_font
                cell_phys.alignment = Alignment(horizontal="center", vertical="center")
                cell_phys.border = border
        
        if col_chimie and len(col_chimie) > 0:
            # Ne fusionner que s'il y a plusieurs colonnes
            if len(col_chimie) > 1 and col_chimie[-1] > col_chimie[0]:
                ws2.merge_cells(start_row=1, start_column=col_chimie[0], end_row=1, end_column=col_chimie[-1])
            for col in range(col_chimie[0], col_chimie[-1]+1):
                cell_chim = ws2.cell(row=1, column=col, value="Chimie" if col==col_chimie[0] else None)
                cell_chim.font = header_font
                cell_chim.alignment = Alignment(horizontal="center", vertical="center")
                cell_chim.border = border
        
        # Noms des salles (ligne 2)
        for idx_s, s in enumerate(salle_list):
            col_letter = ws2.cell(row=2, column=2+idx_s).column_letter
            cell = ws2.cell(row=2, column=2+idx_s, value=s)
            ws2.column_dimensions[col_letter].width = 20
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Date (ligne 2, colonne 1)
        ws2.cell(row=2, column=1, value=formatted_date)
        ws2.cell(row=2, column=1).font = bold_font
        ws2.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row=2, column=1).border = border
        
        # Version simplifiée pour l'affichage (seulement enseignants)
        cell_matrix2 = [[{"content": "", "merge": False, "merge_len": 1, "matiere": ""} for _ in salle_list] for _ in horaires]
        
        for i, c in enumerate(cours):
            # Find assigned room
            salle_assignee = None
            
            # Vérifier si on a une assignation personnalisée pour ce cours (même logique que Planning_Techniciens)
            course_id = c.get('id')
            if custom_room_assignments and course_id:
                try:
                    # S'assurer que custom_room_assignments est un dictionnaire
                    if isinstance(custom_room_assignments, str):
                        import json
                        custom_room_assignments = json.loads(custom_room_assignments)
                    
                    if isinstance(custom_room_assignments, dict):
                        assigned_room_name = custom_room_assignments.get(str(course_id))
                        if assigned_room_name:
                            # Trouver la clé de la salle à partir du nom
                            for room_key, room_data in salles.items():
                                if room_data.get('nom') == assigned_room_name or room_key == assigned_room_name:
                                    salle_assignee = room_key
                                    print(f"🎯 [Affichage] Assignation personnalisée: cours {course_id} -> salle {room_key}")
                                    break
                except Exception as e:
                    print(f"⚠️ [Affichage] Erreur traitement assignations personnalisées: {e}")
                    # En cas d'erreur, continuer avec l'assignation normale
            
            # Si pas d'assignation personnalisée, utiliser l'assignation du solver
            if not salle_assignee:
                for s in salle_list:
                    if (i, s) in x and solver.Value(x[(i, s)]) == 1:
                        salle_assignee = s
                        break
            
            if salle_assignee:
                horaire_debut = c.get('horaire', '8:00')
                start_time = h_to_min(horaire_debut)
                duree = c.get('duree', 110)
                
                # Find start and end indices in horaires (même logique)
                idx_start = 0
                idx_end = 1
                
                for idx, h in enumerate(horaires):
                    if h <= start_time:
                        idx_start = idx
                    if h >= start_time + duree and idx_end == 1:
                        idx_end = idx
                        break
                else:
                    idx_end = len(horaires)
                
                if idx_end <= idx_start:
                    idx_end = idx_start + 1
                
                merge_len = idx_end - idx_start
                
                # Build course content for Affichage (seulement enseignant + niveau) 
                content_affichage = f"{c.get('enseignant', '')}\n{c.get('niveau', '')}"
                
                # Fill matrix avec vérifications d'indices
                try:
                    salle_idx = salle_list.index(salle_assignee)
                except ValueError:
                    continue  # Salle non trouvée, ignorer
                
                # Vérifier que les indices sont dans les limites
                if (idx_start >= 0 and idx_start < len(cell_matrix2) and 
                    salle_idx >= 0 and salle_idx < len(salle_list)):
                    
                    # Remplir seulement la première cellule pour éviter les conflits avec les cellules fusionnées
                    if idx_start < len(cell_matrix2) and salle_idx < len(cell_matrix2[idx_start]):
                        # Utiliser le contenu d'affichage (seulement enseignant + niveau)
                        cell_matrix2[idx_start][salle_idx]["content"] = content_affichage
                        cell_matrix2[idx_start][salle_idx]["matiere"] = c.get('matiere', 'mixte')
                        
                        # Marquer les autres cellules comme "occupées" mais sans contenu
                        for idx in range(idx_start + 1, min(idx_end, len(cell_matrix2))):
                            if idx < len(cell_matrix2) and salle_idx < len(cell_matrix2[idx]):
                                cell_matrix2[idx][salle_idx]["content"] = ""  # Cellule occupée mais vide
                                cell_matrix2[idx][salle_idx]["matiere"] = c.get('matiere', 'mixte')
                    
                    # Marquer pour fusion seulement si valide
                    if merge_len > 1 and idx_start < len(cell_matrix2):
                        cell_matrix2[idx_start][salle_idx]["merge"] = True
                        cell_matrix2[idx_start][salle_idx]["merge_len"] = merge_len
        
        # Remplissage feuille 2 - même logique séparée
        # D'abord, faire toutes les fusions
        merged_ranges2 = set()
        for idx_h, h in enumerate(horaires):
            for idx_s, s in enumerate(salle_list):
                cell = cell_matrix2[idx_h][idx_s]
                excel_row = 3 + idx_h
                excel_col = 2 + idx_s
                
                # Fusion des cellules - seulement pour les cellules de début avec contenu
                if cell["merge"] and cell["content"] and cell["merge_len"] > 1:
                    end_row = excel_row + cell["merge_len"] - 1
                    merge_key = (excel_row, excel_col, end_row)
                    
                    # Vérifier que la fusion est valide et pas déjà faite
                    if (end_row > excel_row and 
                        end_row <= 3 + len(horaires) - 1 and 
                        merge_key not in merged_ranges2):
                        
                        try:
                            ws2.merge_cells(
                                start_row=excel_row, start_column=excel_col,
                                end_row=end_row, end_column=excel_col
                            )
                            merged_ranges2.add(merge_key)
                        except Exception as e:
                            print(f"Erreur fusion feuille 2: start_row={excel_row}, end_row={end_row}, col={excel_col}: {e}")
        
        # Ensuite, remplir le contenu
        for idx_h, h in enumerate(horaires):
            ws2.cell(row=3+idx_h, column=1, value=horaires_str[idx_h])
            ws2.cell(row=3+idx_h, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws2.cell(row=3+idx_h, column=1).border = border
            
            for idx_s, s in enumerate(salle_list):
                cell = cell_matrix2[idx_h][idx_s]
                excel_row = 3 + idx_h
                excel_col = 2 + idx_s
                
                # Vérifier si cette cellule est une cellule fusionnée secondaire
                is_merged_secondary = False
                for merge_key in merged_ranges2:
                    start_row, start_col, end_row = merge_key
                    if (start_col == excel_col and 
                        start_row < excel_row <= end_row):
                        is_merged_secondary = True
                        break
                
                # Appliquer les bordures à toutes les cellules (y compris fusionnées)
                try:
                    ws2.cell(row=excel_row, column=excel_col).border = border
                except Exception:
                    pass
                
                # Remplir seulement les cellules principales (pas les secondaires)
                if not is_merged_secondary:
                    try:
                        ws2.cell(row=excel_row, column=excel_col, value=cell["content"])
                        ws2.cell(row=excel_row, column=excel_col).alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )
                        
                        if cell["content"]:
                            ws2.cell(row=excel_row, column=excel_col).fill = PatternFill(
                                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                            )
                    except Exception as e:
                        # Ignorer les erreurs sur les cellules fusionnées
                        print(f"Ignoring cell write error feuille 2 at ({excel_row}, {excel_col}): {e}")
                        pass
        
        # Hauteur des lignes feuille 2
        for idx_h in range(len(horaires)):
            ws2.row_dimensions[3 + idx_h].height = 20
        
        # Ajout des cours non assignés en bas
        if unassigned_courses:
            # Ligne vide
            ws1.cell(row=3 + len(horaires) + 1, column=1, value="COURS NON ASSIGNÉS")
            ws1.cell(row=3 + len(horaires) + 1, column=1).font = Font(bold=True, color='FF0000', size=14)
            
            row_start = 3 + len(horaires) + 2
            for i, course_info in enumerate(unassigned_courses):
                ws1.cell(row=row_start + i, column=1, value=course_info)
                ws1.cell(row=row_start + i, column=1).font = Font(color='FF0000')
        
        # Sauvegarde du fichier
        from datetime import datetime
        filename = f"planning_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(filename)
        
        assigned_count = len(cours) - len(unassigned_courses)
        print(f"Planning généré: {filename} ({assigned_count}/{len(cours)} cours assignés)")
        return True, filename
        
    except Exception as e:
        return False, f"Erreur lors de la génération Excel: {str(e)}"

def generer_planning_excel(date, end_date=None, return_data_only=False, custom_room_assignments=None):
    """Generate planning Excel file for a specific date or date range"""
    try:
        # Get data from database  
        date_str = date if isinstance(date, str) else date.strftime('%Y-%m-%d')
        raw_requests = database.get_planning_data(date_str)
        raw_rooms = database.get_all_rooms()
        
        # Si on a des assignations personnalisées, modifier les données avant génération
        if custom_room_assignments:
            print(f"📝 Génération avec assignations personnalisées: {custom_room_assignments}")
            # On va continuer la génération normale mais modifier les assignations à la fin
        
        # Récupérer les disponibilités C21
        c21_slots = database.get_c21_availability()
        
        if not raw_requests:
            return False, "Aucune demande trouvée pour cette date"
        
        if not raw_rooms:
            return False, "Aucune salle trouvée dans la base de données"
        
        # Convert rooms to dict format
        salles = {}
        salle_list = []
        for raw_room in raw_rooms:
            # Convert sqlite3.Row to dict
            room = dict(raw_room) if hasattr(raw_room, 'keys') else raw_room
            
            room_name = room.get('name', f'Room_{len(salle_list)}')
            salle_list.append(room_name)
            salles[room_name] = {
                "nom": room_name,
                "type": room.get('type', 'mixte'),
                "ordinateurs": room.get('ordinateurs', 0) or 0,
                "chaises": room.get('chaises', 20) or 20,
                "eviers": room.get('eviers', 0) or 0,
                "hotte": room.get('hotte', 0) or 0,
                "bancs_optiques": room.get('bancs_optiques', 0) or 0,
                "oscilloscopes": room.get('oscilloscopes', 0) or 0,
                "becs_electriques": room.get('becs_electriques', 0) or 0,
                "support_filtration": room.get('support_filtration', 0) or 0,
                "imprimante": room.get('imprimante', 0) or 0,
                "examen": room.get('examen', 0) or 0
            }
        
        # Déterminer le jour de la semaine à partir de la date
        from datetime import datetime
        if isinstance(date, str):
            date_obj = datetime.strptime(date, '%Y-%m-%d')
        else:
            date_obj = date
        
        jours_semaine = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi', 'dimanche']
        jour_planning = jours_semaine[date_obj.weekday()]
        
        # Convert requests to course format
        cours = []
        for i, raw_req in enumerate(raw_requests):
            req = to_dict_request(raw_req)
            material_needs = extract_material_needs(req.get('selected_materials', ''))
            matiere = "mixte"
            if req.get('room_type') == 'Physique':
                matiere = "physique"
            elif req.get('room_type') == 'Chimie':
                matiere = "chimie"
            elif req.get('room_type') == 'Mixte':
                materials_text = str(req.get('selected_materials', '')).lower()
                description_text = str(req.get('material_description', '')).lower()
                combined_text = f"{materials_text} {description_text}"
                physics_keywords = ['oscilloscope', 'générateur', 'signal', 'optique', 'laser', 'prisme', 'lentille', 'physique', 'électricité', 'mécanique', 'ondes']
                chemistry_keywords = ['burette', 'erlenmeyer', 'bécher', 'pipette', 'solution', 'naoh', 'hcl', 'acide', 'base', 'dosage', 'titrage', 'chimie', 'réaction', 'molécule', 'ion', 'ph']
                physics_score = sum(1 for kw in physics_keywords if kw in combined_text)
                chemistry_score = sum(1 for kw in chemistry_keywords if kw in combined_text)
                if physics_score > chemistry_score and physics_score > 0:
                    matiere = "physique"
                elif chemistry_score > physics_score and chemistry_score > 0:
                    matiere = "chimie"
            computers_needed = req.get('computers_needed', 0)
            if computers_needed and computers_needed > 0:
                material_needs["ordinateurs"] = max(material_needs["ordinateurs"], computers_needed)
            cours.append({
                "id": f"{req.get('teacher_name', 'Unknown')}_{i}",
                "enseignant": req.get('teacher_name', 'Unknown'),
                "horaire": req.get('horaire', '9h00') or '9h00',
                "niveau": req.get('class_name', ''),
                "matiere": matiere,
                "jour": jour_planning,
                "ordinateurs": material_needs["ordinateurs"],
                "eviers": material_needs["eviers"],
                "hotte": material_needs["hotte"],
                "bancs_optiques": material_needs["bancs_optiques"],
                "oscilloscopes": material_needs["oscilloscopes"],
                "becs_electriques": material_needs["becs_electriques"],
                "support_filtration": material_needs["support_filtration"],
                "imprimante": material_needs["imprimante"],
                "examen": material_needs["examen"],
                "duree": duree_par_niveau(req.get('class_name', '')),
                "chaises": eleves_par_niveau(req.get('class_name', ''), req.get('teacher_name', 'Unknown')),
                "materiel_demande": req.get('material_description', 'N/A'),
                "selected_materials": req.get('selected_materials', ''),
                "request_name": req.get('request_name', '')
            })

        if not cours:
            return False, "Aucun cours valide à planifier"
        
        # OR-Tools optimization model
        model = cp_model.CpModel()
        x = {}
        poids_salle = {}
        
        # Variables for room assignment
        for i, c in enumerate(cours):
            poids_salle[i] = {}
            for s in salles:
                if compatible(salles[s], c, c21_slots):
                    x[(i,s)] = model.NewBoolVar(f"x_{i}_{s}")
                    
                    # Check if this is a theoretical course (no specific equipment needs)
                    has_equipment_needs = (c["ordinateurs"] > 0 or c["eviers"] > 0 or 
                                          c["hotte"] > 0 or c["bancs_optiques"] > 0 or 
                                          c["oscilloscopes"] > 0 or c["becs_electriques"] > 0 or 
                                          c["support_filtration"] > 0 or c["imprimante"] > 0 or 
                                          c["examen"] > 0)
                    
                    # Weight based on room specialization and REAL EQUIPMENT
                    # AMÉLIORATION: Priorité basée sur les équipements réels, pas seulement le type déclaré
                    
                    # Analyser les équipements réels de la salle
                    room_has_chemistry_equipment = (salles[s]["eviers"] > 0 or salles[s]["hotte"] > 0 or 
                                                   salles[s]["becs_electriques"] > 0 or salles[s]["support_filtration"] > 0)
                    room_has_physics_equipment = (salles[s]["oscilloscopes"] > 0 or salles[s]["bancs_optiques"] > 0)
                    
                    # Pondération intelligente basée sur matière + équipements réels
                    if c["matiere"] == "chimie":
                        if room_has_chemistry_equipment:
                            # Salle avec équipements de chimie (C22, C24, C31-33) - PRIORITÉ MAXIMALE
                            poids_salle[i][s] = 10 if has_equipment_needs else 9
                        elif salles[s]["type"] == "chimie":
                            # Salle déclarée chimie mais sans équipements - BONNE
                            poids_salle[i][s] = 8 if has_equipment_needs else 7
                        elif salles[s]["type"] == "mixte" and not room_has_physics_equipment:
                            # Salle mixte sans équipements physique - ACCEPTABLE
                            poids_salle[i][s] = 6
                        elif room_has_physics_equipment or salles[s]["type"] == "physique":
                            # Salle avec équipements physique - À ÉVITER
                            poids_salle[i][s] = 2
                        else:
                            # Dernière option
                            poids_salle[i][s] = 4
                            
                    elif c["matiere"] == "physique":
                        if room_has_physics_equipment:
                            # Salle avec équipements de physique (C25, C27) - PRIORITÉ MAXIMALE
                            poids_salle[i][s] = 10 if has_equipment_needs else 9
                        elif salles[s]["type"] == "physique":
                            # Salle déclarée physique mais sans équipements - BONNE
                            poids_salle[i][s] = 8 if has_equipment_needs else 7
                        elif salles[s]["type"] == "mixte" and not room_has_chemistry_equipment:
                            # Salle mixte sans équipements chimie - ACCEPTABLE
                            poids_salle[i][s] = 6
                        elif room_has_chemistry_equipment or salles[s]["type"] == "chimie":
                            # Salle avec équipements chimie - À ÉVITER
                            poids_salle[i][s] = 2
                        else:
                            # Dernière option
                            poids_salle[i][s] = 4
                            
                    else:  # matiere == "mixte" ou autres
                        if salles[s]["type"] == "mixte":
                            poids_salle[i][s] = 7  # Parfait pour cours mixtes
                        elif not room_has_chemistry_equipment and not room_has_physics_equipment:
                            poids_salle[i][s] = 6  # Salles théoriques sont bien
                        else:
                            poids_salle[i][s] = 4  # Éviter les salles spécialisées pour cours mixtes
                else:
                    poids_salle[i][s] = 0

        # Constraints: each course with compatible rooms must have exactly one room
        for i in range(len(cours)):
            compatible_rooms = [x[(i,s)] for s in salles if (i,s) in x]
            if compatible_rooms:
                model.Add(sum(compatible_rooms) == 1)
            else:
                print(f"⚠️ ATTENTION: Cours {i} ({cours[i]['enseignant']} - {cours[i]['niveau']}) n'a AUCUNE salle compatible!")
                print(f"   Matière: {cours[i]['matiere']}, Horaire: {cours[i]['horaire']}")
                print(f"   Besoins: ordinateurs={cours[i]['ordinateurs']}, eviers={cours[i]['eviers']}, hotte={cours[i]['hotte']}")

        # Constraints: no room conflicts (time overlap)
        for i in range(len(cours)):
            for j in range(i + 1, len(cours)):
                start1, end1 = interval_cours(cours[i])
                start2, end2 = interval_cours(cours[j])
                
                # Check if courses overlap in time
                if not (end1 <= start2 or end2 <= start1):
                    # They overlap, so they can't use the same room
                    for s in salles:
                        if (i,s) in x and (j,s) in x:
                            model.Add(x[(i,s)] + x[(j,s)] <= 1)

        # Get list of teachers
        enseignants = list(set(c["enseignant"] for c in cours))
        
        # Preference for same teacher to use same room
        objectif_pref = []
        for enseignant in enseignants:
            cours_ens = [i for i, c in enumerate(cours) if c["enseignant"] == enseignant]
            for idx1 in range(len(cours_ens)):
                for idx2 in range(idx1 + 1, len(cours_ens)):
                    i, j = cours_ens[idx1], cours_ens[idx2]
                    for s in salle_list:
                        if (i, s) in x and (j, s) in x:
                            pref = model.NewBoolVar(f"pref_{i}_{j}_{s}")
                            model.AddBoolAnd([x[(i, s)], x[(j, s)]]).OnlyEnforceIf(pref)
                            model.AddBoolOr([x[(i, s)].Not(), x[(j, s)].Not()]).OnlyEnforceIf(pref.Not())
                            objectif_pref.append(pref)

        # Variables for room usage
        salle_utilisee = {}
        for s in salles:
            salle_utilisee[s] = model.NewBoolVar(f"salle_utilisee_{s}")
            vars_for_salle = [x[(i,s)] for i in range(len(cours)) if (i,s) in x]
            if vars_for_salle:
                model.AddMaxEquality(salle_utilisee[s], vars_for_salle)
            else:
                # Aucune variable pour cette salle -> elle n'est pas utilisée
                model.Add(salle_utilisee[s] == 0)

        # Objective: maximize room specialization + teacher preference + room usage
        model.Maximize(
            sum(x[(i,s)] * poids_salle[i][s] for (i,s) in x) +  # room specialization
            sum(objectif_pref) +  # same room for same teacher
            0.1 * sum(salle_utilisee.values())  # encourage room usage diversity
        )
        
        # Solve the model
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 60
        status = solver.Solve(model)

        print(f"Statut de la résolution: {status}")
        print(f"Nombre de cours: {len(cours)}")
        print(f"Nombre de salles disponibles: {len(salles)}")
        print(f"Salles: {list(salles.keys())}")
        print(f"Variables x créées: {len(x)}")
        
        # Debug: afficher les cours et leurs salles compatibles
        for i, c in enumerate(cours):
            compatible_rooms_list = [s for s in salles if (i,s) in x]
            print(f"Cours {i} ({c['enseignant']} - {c['niveau']} à {c['horaire']}): {len(compatible_rooms_list)} salles compatibles - {compatible_rooms_list}")
        
        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            # Count assignments
            assignments = 0
            unassigned_courses = []
            for i, c in enumerate(cours):
                assigned = False
                for s in salles:
                    if (i, s) in x and solver.Value(x[(i, s)]) == 1:
                        assignments += 1
                        assigned = True
                        break
                if not assigned:
                    unassigned_courses.append(f"{c['enseignant']} - {c['niveau']} à {c['horaire']}")
            
            print(f"Cours assignés: {assignments}/{len(cours)}")
            if unassigned_courses:
                print("Cours NON assignés:")
                for course in unassigned_courses:
                    print(f"  - {course}")
            
            # Generate Excel even with partial assignments
            if return_data_only:
                # Retourner les données structurées pour l'éditeur
                courses_data = []
                assignments_data = {}
                room_assignments = {}
                
                for i, course in enumerate(cours):
                    course_id = course['id']  # Utiliser l'ID artificiel généré par la conversion 
                    courses_data.append({
                        'id': course_id,
                        'subject': course['matiere'],
                        'level': course['niveau'],
                        'teacher': course['enseignant'],
                        'date': date_str,  # Utiliser date_str de la fonction
                        'time': course['horaire'],
                        'duration': course['duree'],
                        'students': course.get('chaises', 20),  # Utiliser 'chaises' au lieu de 'nb_eleves'
                        'request_name': course.get('request_name', ''),
                        'material_description': course.get('materiel_demande', ''),
                        'ordinateurs': course.get('ordinateurs', 0),
                        'eviers': course.get('eviers', 0),
                        'hotte': course.get('hotte', 0),
                        'bancs_optiques': course.get('bancs_optiques', 0),
                        'oscilloscopes': course.get('oscilloscopes', 0),
                        'becs_electriques': course.get('becs_electriques', 0),
                        'support_filtration': course.get('support_filtration', 0),
                        'imprimante': course.get('imprimante', 0),
                        'examen': course.get('examen', 0)
                    })
                    
                    # Trouver l'assignation de salle
                    assigned_room = None
                    for s in salles:
                        if (i, s) in x and solver.Value(x[(i, s)]) == 1:
                            assigned_room = s
                            course_id = course['id']
                            print(f"Cours {course_id} ({course['enseignant']} - {course['niveau']}) assigné à {s} ({salles[s]['nom']})")
                            slot_key = f"{date_str}_{course['horaire']}"
                            if slot_key not in assignments_data:
                                assignments_data[slot_key] = []
                            assignments_data[slot_key].append(course_id)
                            break
                    
                    # Ajouter l'assignation de salle pour l'interface
                    course_id = course['id']
                    if assigned_room:
                        room_assignments[course_id] = salles[assigned_room]['nom']
                    else:
                        room_assignments[course_id] = 'Non assigné'
                
                # Créer la liste des jours
                if end_date:
                    days = []
                    start = datetime.strptime(date, '%Y-%m-%d').date() if isinstance(date, str) else date
                    end = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
                    current_date = start
                    while current_date <= end:
                        if current_date.weekday() < 5:  # Lundi à vendredi
                            days.append(current_date.strftime('%Y-%m-%d'))
                        current_date += timedelta(days=1)
                else:
                    days = [date if isinstance(date, str) else date.strftime('%Y-%m-%d')]
                
                # Horaires autorisés pour le début des cours (pas toutes les 15 minutes)
                time_slots = [
                    '9h00', '9h30', '10h00', '10h45', '11h15', '11h45', '12h15', '12h45',
                    '13h15', '13h45', '14h15', '14h45', '15h15', '15h45', '16h15', '16h45', '17h15'
                ]
                
                print(f"⏰ Créneaux autorisés pour début de cours - Total: {len(time_slots)} horaires")
                print(f"📝 Tous les créneaux: {time_slots}")
                
                rooms_list = [{'id': s, 'name': salles[s]['nom']} for s in salles]
                print(f"📍 Rooms list pour l'API: {rooms_list}")
                print(f"📍 Room assignments: {dict(list(room_assignments.items())[:5])}")  # Premiers 5 pour debug
                
                return True, {
                    'courses': courses_data,
                    'days': days,
                    'time_slots': time_slots,
                    'assignments': assignments_data,
                    'room_assignments': room_assignments,
                    'rooms': rooms_list
                }
            else:
                return generer_excel_optimise(cours, salles, x, solver, unassigned_courses, date_str, custom_room_assignments)
        
        elif status == cp_model.INFEASIBLE:
            return False, "ERREUR: Il y a plus de cours simultanés que de salles disponibles! Impossible de générer le planning."
        
        else:
            return False, f"ERREUR: Résolution échouée avec le statut: {status}. Impossible de générer le planning."

    except Exception as e:
        import traceback
        return False, f"Erreur lors de la génération du planning: {str(e)}\n{traceback.format_exc()}"


def get_planning_data_for_editor(target_date):
    """
    Appelle la nouvelle version qui fonctionne
    """
    return get_planning_data_for_editor_v2(target_date)


def get_planning_data_for_editor_v2(target_date):
    """
    Version qui utilise directement la génération normale avec return_data_only=True
    mais en corrigeant le problème de clés.
    """
    try:
        print(f"🔧 [EDITOR] Début génération du planning - Date: {target_date}")
        
        # Retourner à la méthode return_data_only=True mais avec les corrections
        print(f"📞 [EDITOR] Appel de generer_planning_excel avec return_data_only=True")
        success, result = generer_planning_excel(target_date, return_data_only=True)
        print(f"📋 [EDITOR] Retour de generer_planning_excel: success={success}")
        
        if not success:
            print(f"❌ Échec de la génération: {result}")
            return {
                'courses': [],
                'days': [target_date],
                'time_slots': ['9h00', '9h30', '10h00', '10h45', '11h15', '11h45', '12h15', '12h45', '13h15', '13h45', '14h15', '14h45', '15h15', '15h45', '16h15', '16h45', '17h15'],
                'assignments': {},
                'rooms': []
            }
        
        print(f"✅ Planning avec assignations généré avec succès")
        return result
        
    except Exception as e:
        print(f"❌ Erreur dans get_planning_data_for_editor_v2: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'courses': [],
            'days': [target_date],
            'time_slots': ['9h00', '9h30', '10h00', '10h45', '11h15', '11h45', '12h15', '12h45', '13h15', '13h45', '14h15', '14h45', '15h15', '15h45', '16h15', '16h45', '17h15'],
            'assignments': {},
            'rooms': []
        }





if __name__ == "__main__":
    # Test the function
    success, message = generer_planning_excel("2024-09-29")
    print(message)